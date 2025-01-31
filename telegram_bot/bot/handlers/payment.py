import datetime
import time
import uuid

from aiogram import Bot
from aiohttp import web
from django.db.models import Sum, F
from yookassa import Configuration, Payment

from ..keyboards import *
from ..texts import *

from main_app.models import UserBucket
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from config import YOOKASSA_ID, YOOKASSA_SECRET

Configuration.account_id = YOOKASSA_ID
Configuration.secret_key = YOOKASSA_SECRET

# excel setup
EXCEL_PATH = f'{Path(__file__).resolve().parent.parent.parent.parent}.logs/заказы.xlsx'

if os.path.exists(EXCEL_PATH):
    workbook = load_workbook(EXCEL_PATH)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    headers = [
        "Дата, время", "ID пользователя", "@username", "Имя Фамилия", "Место доставки", "Общая цена",
        'Название', 'Количество', 'Цена'
    ]
    sheet.append(headers)

sheet = workbook.active
for column_letter in 'ABCDEFGHI':
    sheet.column_dimensions[column_letter].auto_size = True


def yookassa_wrapper(bot: Bot):
    async def yookassa_webhook(request):
        data = (await request.json())['object']

        place = data['metadata']['place']
        user_id = data['metadata']['user_id']

        user = await User.objects.aget(id=user_id)

        if data['status'] == 'succeeded':
            await bot.send_message(
                chat_id=user_id,
                text=get_payment_succeed_text(user=user, place=place),
                reply_markup=get_menu_keyboard()
            )
            now_time = datetime.datetime.now()
            sheet.append([
                now_time.strftime('%Y-%m-%d %H:%M'),
                user_id, user.username, f'{user.first_name} {user.last_name}',
                place, f'{await get_bucket_cost(user.id)}руб.'
            ])

            data_to_excel = list()

            user_bucket = await sync_to_async(UserBucket.objects.filter)(user=user)
            async for item in user_bucket:
                goods = await sync_to_async(lambda: item.goods)()
                data_to_excel.append([goods.name, item.amount, f'{goods.cost * item.amount}руб'])
                await item.adelete()

            max_row = sheet.max_row + 1
            for row, data in enumerate(data_to_excel, max_row):
                for i, col in enumerate(range(7, len(data) + 7)):
                    sheet.cell(row, col, data[i])

            workbook.save(EXCEL_PATH)

        elif data['status'] == 'canceled':
            await bot.send_message(
                chat_id=user_id,
                text=get_payment_canceled_text(user=user),
                reply_markup=get_menu_keyboard()
            )

        return web.Response(status=200)

    return yookassa_webhook


async def get_bucket_cost(user_id: int):
    return await sync_to_async(
        lambda: UserBucket.objects.filter(user_id=user_id)
        .annotate(total_cost=F('goods__cost') * F('amount'))
        .aggregate(total_sum=Sum('total_cost'))['total_sum']
    )()


async def get_link(user: User, place: str, bot: Bot) -> str:
    bucket_cost = await get_bucket_cost(user.id)

    bot_info = await bot.get_me()
    payment = Payment.create({
        "amount": {
            "value": f"{bucket_cost}.00",
            "currency": "RUB"
        },
        'metadata': {
            'user_id': user.id,
            'place': place
        },
        "confirmation": {
            "type": "redirect",
            "return_url": f"{bot_info.url}"
        },
        "capture": True,
        "description": f"Оплата товара через бота @{bot_info.username}"
    }, uuid.uuid4())

    return payment.confirmation['confirmation_url']
